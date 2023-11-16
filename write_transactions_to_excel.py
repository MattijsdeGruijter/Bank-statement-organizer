import openpyxl
import tkinter as tk
import sys
from openpyxl.styles import PatternFill, Font

class WriteTransactionsToExcel:
    def __init__(self, filename, json_dict_obj, totals_month, transactions_per_month_per_category) -> None:
        self.filename = filename
        self.data_only = False
        self.json_dict_obj = json_dict_obj
        self.totals_month = totals_month
        self.transactions_per_month_per_category = transactions_per_month_per_category
        self.fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type="solid")

    def open_workbook(self):
        '''
        Tries to open the workbook, gives an error if it
        fails to open the file
        '''
        try:
            self.workbook = openpyxl.load_workbook(self.filename, data_only=self.data_only)
            return self.workbook
        except PermissionError:
            # Display pop-up message box
            root = tk.Tk()
            root.withdraw()
            tk.messagebox.showinfo("Permission Denied", "Please close the workbook '{}' and try again.".format(self.filename))
            return sys.exit()
    
    def write_totals_per_category_to_total_sheet(self) -> None:
        '''
        Creates a new sheet in the workbook called 'Total'
        if there is already a 'Total' sheet, it will be
        removed and recreated.
        
        Every category will have his own row, where the 
        totals of each month will be written in columns.
        '''
        self.open_workbook()
        self.total_sheet__create_total_sheet()
        self.total_sheet__put_categories_in_first_column_and_format()
        self.total_sheet__write_totals_per_month_in_the_next_columns()
        self.save_workbook()
    
    def total_sheet__create_total_sheet(self) -> None:
        if 'Total' in self.workbook.sheetnames:
            total_sheet = self.workbook['Total']
            self.workbook.remove(total_sheet)
        self.sheet_total = self.workbook.create_sheet('Total')
        self.sheet_total.column_dimensions['A'].width = 20
        
    def total_sheet__put_categories_in_first_column_and_format(self) -> None:       
        '''
        sort the dict so that the overall sum of 
        the most expencive category is on top
        write the categories to the first column
        also format the even rows with a grey color
        '''
        row_num = 2
        self.totals = self.save_load.data
        totals_sorted = dict(sorted(self.totals.items(), key=lambda x: abs(x[1]['total']), reverse=True))
        for cat in totals_sorted:
            self.sheet_total.cell(row=row_num, column=1).value = cat
            if row_num % 2 == 0:
                self.sheet_total.cell(row=row_num, column=1).fill = self.fill
            row_num += 1
        
    def total_sheet__write_totals_per_month_in_the_next_columns(self):
        '''
        pretty self explanitory i think
        '''
        for month, category_dict in self.totals_month.items():
            
            # Find the first empty column for the date and amounts
            col_num = 2
            while self.sheet_total.cell(row=1, column=col_num).value:
                col_num += 1
            # Write the date to cell 
            self.sheet_total.cell(row=1, column=col_num).value = month
            self.sheet_total.cell(row=1, column=col_num).font = Font(bold=True)
                    
            #Go through all the categories and check against the category in the dict. If it matches, write the value
            for cat_cell in self.sheet_total.iter_rows(min_row=2, max_row=len(self.totals)+2, min_col=1):  
                if cat_cell[0].row % 2 == 0:
                    self.sheet_total.cell(row=cat_cell[0].row, column=col_num).fill = self.fill
                for category, bedrag in category_dict.items():
                    if cat_cell[0].value == category:
                        self.sheet_total.cell(row=cat_cell[0].row, column=col_num).value = bedrag
                        if cat_cell[0].row % 2 == 0:
                            self.sheet_total.cell(row=cat_cell[0].row, column=col_num).fill = self.fill
       
    def write_month_transactions(self, transactions_per_month_per_category:dict) -> None:
        '''
        create excelsheets for all the months, and writes the categories and theri totals.
        Then calls the write_column_transactions function to fill out all the transactions per category
        '''
        self.open_workbook()
        
        #per month, look at the categories for that month
        for month , transactions_per_category in transactions_per_month_per_category.items():
            sheet_month = self.other_sheets__create_sheet_and_format_cells(month)
            sorted_transactions_per_category = dict(sorted(transactions_per_category.items(), key=lambda x: abs(x[1]["Totaal"]), reverse=True))
    
            #per category, check the transactions
            for category, transaction_dict in sorted_transactions_per_category.items():
                self.other_sheets__type_check(category, transaction_dict, sheet_month)
        
        self.save_workbook()

    def other_sheets__create_sheet_and_format_cells(self, month):
        '''
        Creates a sheet for the given month
        Formats the cells accordingly and puts some basic info like start and end saldo
        returns the newly created sheet
        '''
        #create empty sheet (and delete the old one if it exists)
        month_string = str(month)
        if month_string in self.workbook.sheetnames:
            month_sheet = self.workbook[month_string]
            self.workbook.remove(month_sheet)
        sheet_month = self.workbook.create_sheet(month_string)
        
        #format cells
        sheet_month.column_dimensions['A'].width = 35
        sheet_month.column_dimensions['E'].width = 35
        sheet_month.column_dimensions['I'].width = 65
        
        
        # get the first date and saldo, and last date and saldo, and the difference between the last saldo and the first saldo of the month.
        list_of_transactions = self.transactions_per_month_per_category[month]        
        first_date = list_of_transactions[-1]
        last_date = list_of_transactions[0]
        first_saldo = first_date['Saldo na mutatie'] - first_date['Bedrag']
        last_saldo = last_date['Saldo na mutatie']
        difference = last_saldo - first_saldo
        
        #write the first and last saldo, and the difference to the sheet  
        sheet_month.cell(row=1, column=1).value = month       
        sheet_month.cell(row=2, column=1).value = 'Start saldo'
        sheet_month.cell(row=3, column=1).value = 'Profit this month'
        sheet_month.cell(row=4, column=1).value = 'End saldo'
        
        sheet_month.cell(row=2, column=2).value = first_saldo
        sheet_month.cell(row=3, column=2).value = difference
        sheet_month.cell(row=4, column=2).value = last_saldo
        
        # make the difference red if the difference is negative
        if difference < 0:
            sheet_month.cell(row=3, column=2).font = Font(color='FF0000')

        return sheet_month
          
    def other_sheets__type_check(self, category, transaction_dict, sheet_month):
        '''
        Checks for each category if it falls into fixed cost, returning cost or exception cost
        Based on this, picks a column and calls the write_column_transactions method to 
        write all the transactions for this category into that column
        '''
        MAIN_CATEGORIE_DICT = self.save_load.data
        row_nr_A = 6
        color_A = 'BFBFBF'
        row_nr_E = 2
        color_B = 'FFD966'
        row_nr_I = 2
        color_C = 'A9D08E'
        json_dict = self.json_dict_obj.data
        type_check = json_dict[category]['type']
        
        # sort the transactions by date
        sorted_data = sorted(((k, v) for k, v in transaction_dict.items() if k != 'Totaal'), key=lambda x: x[1]['Datum'], reverse=False)
        sorted_transaction_dict = dict(sorted_data)
        sorted_transaction_dict['Totaal'] = transaction_dict['Totaal']
        
        #determain the column to write the transaction to, based on the type of transaction
        if  type_check== "fixed cost":
            row_nr_A = self.otehr_sheets__write_column_transactions(category, sheet_month, sorted_transaction_dict, row_nr_A, col_nr=1, color=color_A)
            row_nr_A += 1
        elif type_check== "returning cost":
            row_nr_E = self.otehr_sheets__write_column_transactions(category, sheet_month, sorted_transaction_dict, row_nr_E, col_nr=5, color=color_B)
            row_nr_E += 1
        elif type_check== "exception cost":
            row_nr_I = self.otehr_sheets__write_column_transactions(category, sheet_month, sorted_transaction_dict, row_nr_I, col_nr=9, color=color_C)
            row_nr_I += 1
        
    def otehr_sheets__write_column_transactions(self, category, sheet_month, transaction_dict, row_nr, col_nr, color) -> int:
        '''
        write the rest all the transactions from the given category in the right column
        returns the last row that was written to
        '''
        #write the category name and the total amount to the first row
        sheet_month.cell(row=row_nr, column=col_nr).value = category
        sheet_month.cell(row=row_nr, column=col_nr+1).value = transaction_dict['Totaal']
        
        # make both values bold
        sheet_month.cell(row=row_nr, column=col_nr).font = Font(bold=True)
        sheet_month.cell(row=row_nr, column=col_nr+1).font = Font(bold=True) 
        # make the amount red if negative
        if transaction_dict['Totaal'] < 0:
            sheet_month.cell(row=row_nr, column=col_nr+1).font = Font(color='FF0000', bold=True)
            
        # give the category name and the total amount a color
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        sheet_month.cell(row=row_nr, column=col_nr).fill = fill
        sheet_month.cell(row=row_nr, column=col_nr+1).fill = fill
        
        #write the the transactions of this category to the rows below
        for transaction, transaction_dict in transaction_dict.items():
            if not transaction == 'Totaal':
                while sheet_month.cell(row=row_nr, column=col_nr).value:
                    row_nr += 1  
                korte_naam = transaction_dict['Korte naam']
                sheet_month.cell(row=row_nr, column=col_nr).value = korte_naam
                sheet_month.cell(row=row_nr, column=col_nr+1).value = transaction_dict['Bedrag']
                if transaction_dict['Bedrag'] < 0:
                    sheet_month.cell(row=row_nr, column=col_nr+1).font = Font(color='FF0000')
                str_day = self.otehr_sheets__convert_date_to_string_day(transaction_dict['Datum'])
                sheet_month.cell(row=row_nr, column=col_nr+2).value = str_day 
        # self.wb_afschriften.save(self.filename)
        return row_nr   
    
    def otehr_sheets__convert_date_to_string_day(self, date) -> str:
        '''
        convertes a date-time date to a day string
        Very sophisticated as you can see
        '''
        str_Year_month_day = str(date)
        str_day = str_Year_month_day[6:]
        if str_day[0] == '0':
            str_day = str_day[-1]
        if str_day[0] == '1' and len(str_day) == 2:
            str_day += 'th'
        else:
            if str_day[-1] == '1':
                str_day += 'st'
            elif str_day[-1] == '2':
                str_day += 'nd'
            elif str_day[-1] == '3':
                str_day += 'rd'
            else:
                str_day += 'th'
        return str_day
                
    def save_workbook(self) -> None:
        self.workbook.save(self.filename)
