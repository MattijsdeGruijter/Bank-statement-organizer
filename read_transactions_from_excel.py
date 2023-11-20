from settings import MAX_ROW
import re
import datetime
import openpyxl
import tkinter as tk
import sys


class ReadTransactionsFromExcel:
    def __init__(self, json_dictionary_manager, filename):
        self.filename = filename
        self.data_only = False
        self.json_dictionary_manager = json_dictionary_manager
        self.transactions_per_month_per_category = {}
        self.totals_per_month = {}
        self.transactions_per_month = {}
    
    def open_workbook(self):
        '''
        Tries to open the workbook, gives an error if it
        fails to open the file
        '''
        try:
            self.workbook = openpyxl.load_workbook(self.filename, data_only=self.data_only)
            return self.workbook
        except PermissionError:
            # Display pop-up message box
            root = tk.Tk()
            root.withdraw()
            tk.messagebox.showinfo("Permission Denied", "Please close the workbook '{}' and try again.".format(self.filename))
            return sys.exit()
       
    def get_transaction(self, date_row):
        '''
        Takes a excel row and returns a transaction in the form of a dictionary
        Also rounds off the bedrag value to 2 digits, and reformats the 
        mededeling and naam_omschrijving to make them more readable.
        
        Also calls fill_dictionaries to categorize transactions and 
        calculates totals
        '''
        date_cell, naam_omschrijving_cell, rekening_cell, tegenrekening_cell, code_cell, af_bij_cell, bedrag_raw_cell, mutatiesoort_cell, mededeling_cell, saldo_na_mutatie_cell, a_cell, s_cell, d_cell = date_row
        date = date_cell.value
        naam_omschrijving = naam_omschrijving_cell.value
        af_bij=  af_bij_cell.value
        bedrag_raw = bedrag_raw_cell.value
        mededeling = mededeling_cell.value
        saldo_na_mutatie = saldo_na_mutatie_cell.value
        
        
        
        bedrag = self.float_precision(bedrag_raw)
        if af_bij == 'Af':
            bedrag = -bedrag
        mededeling_new, naam_omschrijving_new = self.rewrite_transaction(mededeling, naam_omschrijving)
        category = self.categorize_transaction(naam_omschrijving_new, mededeling_new, bedrag)
        if category == None:
            print(category)
        transaction = {
            'Datum': date,
            'Bedrag': self.float_precision(bedrag),
            'Naam': mededeling,
            'Category': category,
            'Korte naam': naam_omschrijving_new,
            'Mededeling': mededeling_new,
            'Saldo na mutatie': saldo_na_mutatie
        }
        
        self.fill_dictionaries(date, transaction, category, bedrag)
        
        # return transaction

    def categorize_transaction(self, naam_omschrijving, mededeling, bedrag):
        '''
        Looks through the json dictionary and goes though the key words
        of each category. If there is a match between the keyword and 
        the mededeling or naam_omschrijving, the bedrag amount of that 
        category is updated in the json dictionary.
        
        Do we really need this function? Seems redundant
        '''
        for cat, cat_dict in self.json_dictionary_manager.data.items():
            if cat != "Other": 
                # if 'betaalverzoek' in naam_omschrijving.lower() or naam_omschrijving.lower()[0:3] == 'hr ' or naam_omschrijving.lower()[0:3] == 'mw ':
                #     category = 'Overschrijving'
                #     pattern = r":(.*?)IBAN:"
                #     match = re.search(pattern, mededeling)
                #     if match:
                #         naam_omschrijving = match.group(1)
                keywords = cat_dict['key_words']
                for keyword in keywords:
                    if keyword in mededeling.lower() or keyword in naam_omschrijving.lower():
                        category = cat
                        self.json_dictionary_manager.data[category]["total"] += self.float_precision(bedrag)  
                        self.json_dictionary_manager.save_dictionary_to_file()
                        return category 
        return "Overschrijving"

            
    
    def rewrite_transaction(self, mededeling, naam_omschrijving):
        """
        Extracts the relevant part of the mededeling/naam_omschrijving.
        Sub function of the get_transaction method
        """
        
    # Define a dictionary to map keywords to new values
        keyword_mappings = {
            "spaarrekening": r"spaarrekening (.*?)Valuta",
            "belasting": r"Omschrijving: (.*?)IBAN",
            "betaalverzoek": r"Omschrijving: (.*?)IBAN",
            "tikkie": r"Naam: (.*?)NL"
        }

        # Extract mededeling_only using the first pattern
        pattern1 = r"Omschrijving:(.*?)IBAN:"
        match1 = re.search(pattern1, mededeling)
        mededeling_only = match1.group(1) if match1 else ""

        # Check for keywords and update naam_omschrijving
        for keyword, pattern in keyword_mappings.items():
            if keyword in mededeling.lower():
                match2 = re.search(pattern, mededeling)
                if match2:
                    naam_omschrijving = match2.group(1)
                if keyword in ["betaalverzoek", "tikkie"]:
                    category = "Overschrijving"
                break  # Exit the loop once a match is found

        return mededeling_only, naam_omschrijving
        
    
    def go_through_excel_file(self):
        '''
        goes through the excel file and feeds every single row into the 
        get_transaction method
        '''
        self.open_workbook()
        all_sheets = self.workbook.sheetnames
        self.sheet = self.workbook[all_sheets[0]]
        for date_row in self.sheet.iter_rows(min_row=2, min_col=1, max_col=13, max_row=MAX_ROW):
            if date_row[0].value is not None:
                self.get_transaction(date_row)
            else:
                break
        return self.transactions_per_month_per_category, self.totals_per_month, self.transactions_per_month
    
    def fill_dictionaries(self, date, transaction, category, bedrag):  
        '''
        Categorizes transactions per month and get totals per month per 
        category. 
        '''       
        date_in_datetime = datetime.datetime.strptime(str(date), '%Y%m%d')
        month = date_in_datetime.strftime('%Y-%m')
        if month not in self.transactions_per_month_per_category:
            self.transactions_per_month_per_category[month] = {}
        
        if month not in self.transactions_per_month:
            self.transactions_per_month[month] = []
        
        self.transactions_per_month[month].append(transaction)
            
        if category not in self.transactions_per_month_per_category[month]:
            self.transactions_per_month_per_category[month][category] = {}
            self.transactions_per_month_per_category[month][category]["Totaal"] = 0    
        self.transactions_per_month_per_category[month][category][transaction["Naam"]] = transaction
        self.transactions_per_month_per_category[month][category]["Totaal"] += bedrag
        
        if month not in self.totals_per_month:
            self.totals_per_month[month] = {}
            
        if category not in self.totals_per_month[month]:
            self.totals_per_month[month][category] = 0
        
        self.totals_per_month[month][category] += bedrag
        self.totals_per_month[month][category] = self.float_precision(self.totals_per_month[month][category])
        
        #sort transactions by bedrag    
        for key, value in self.transactions_per_month_per_category[month].items():
            if isinstance(value, dict) and 'Totaal' not in value:
                sorted_nested = dict(sorted(value.items(), key=lambda x: abs(x[1]['Bedrag']), reverse=True))
                self.transactions_per_month_per_category[key] = sorted_nested

        # for key, value in self.transactions_per_month[month].items():
        #     if isinstance(value, dict) and 'Totaal' not in value:
        #         sorted_nested2 = dict(sorted(value.items(), key=lambda x: abs(x[1]['Bedrag']), reverse=True))
        #         self.transactions_per_month[key] = sorted_nested2

        

    
    #deals with float precision
    def float_precision(self, input):
        if input is not None:
            input_rounded = "{:.2f}".format(input)
            output = float(input_rounded)
            return output